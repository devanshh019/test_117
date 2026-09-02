# Excel Spreadsheet (.xlsx) Strategy Renderer
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .base import Renderer
from ..schemas import XlsxSpec, DocumentResult, DocType


class XlsxRenderer(Renderer):

    def render(self, spec: XlsxSpec, output_path: Path) -> DocumentResult:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = spec.sheet_name[:30]

        header_hex = "{:02X}{:02X}{:02X}".format(*self.theme.rgb_dark)
        for col_num, h in enumerate(spec.headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(spec.rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                self._apply_rules(cell, val, spec.rules)

        wb.save(output_path)

        return DocumentResult(
            filename=output_path.name,
            path=str(output_path),
            doc_type=DocType.XLSX,
            size_bytes=os.path.getsize(output_path),
        )

    def _apply_rules(self, cell, val, rules) -> None:
        """Applies solid cell fill highlighting dynamically based on rule.color_hex."""
        val_str = str(val).strip().upper()
        for rule in rules:
            if any(m.upper() == val_str or m.upper() in val_str.split() for m in rule.match_values):
                color = str(rule.color_hex or "{:02X}{:02X}{:02X}".format(*self.theme.rgb_orange)).lstrip("#").upper()
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=rule.bold, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                return
